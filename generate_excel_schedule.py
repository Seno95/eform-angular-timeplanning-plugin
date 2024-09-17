import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# list for the dates
start_date = datetime.today()
dates = [start_date + timedelta(days=i) for i in range(10)]

# plan timer values from 1 to 10
plan_timer = list(range(1, 11))

# Create plan tekst values (time ranges based on Plan timer)
def create_time_range(plan_timer):
    start_time = datetime.strptime("08:00", "%H:%M")
    end_time = start_time + timedelta(hours=plan_timer)
    return f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}"

plan_tekst = [create_time_range(i) for i in plan_timer]

# Employee information
employee_name = ["John Doe"] * 10
employee_id = [1] * 10  # Setting the ID as "1" for all rows

# week numbers and weekdays
week_numbers = [(date.isocalendar()[1]) for date in dates]  # ISO week number for each date
week_days = [(date.strftime('%A')) for date in dates]  # Day of the week for each date

# Create a dictionary for the DataFrame
data = {
    "Dato": [date.strftime('%d.%m.%Y') for date in dates],
    "Plan timer": plan_timer,
    "Plan tekst": plan_tekst,
    "Medarbejder": employee_name,
    "Medarbejder ID": employee_id,  # New column for the employee ID
    "Uge": week_numbers,
    "Ugedag": week_days
}

# Create  DataFrame
df = pd.DataFrame(data)

# output directory and filename
output_dir = 'eform-client/cypress/fixtures/'
output_file = 'output_schedule.xlsx'

# Check if directory exist
if not os.path.exists(output_dir):
    raise FileNotFoundError(f"The directory {output_dir} does not exist. Please ensure the directory is present.")

# Save  file to directory
file_path = os.path.join(output_dir, output_file)
df.to_excel(file_path, index=False)

# Load  workbook to apply formatting
wb = load_workbook(file_path)
ws = wb.active

# header font bold + large
header_font = Font(size=12, bold=True)

# font header row
for col in range(1, ws.max_column + 1):
    ws.cell(row=1, column=col).font = header_font

# column width
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[col_letter].width = adjusted_width

# formatting changes
wb.save(file_path)

print(f"Excel file created and formatted at: {file_path}")
